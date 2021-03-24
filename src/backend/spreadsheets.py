import datetime
import logging

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import PatternFill, Font, Fill, Color
from openpyxl.utils import get_column_letter

from src.backend import constants as cnst
from src.backend import data_calls as dc
from src.backend import data


#   returns an list of dictionaries
def sorted_dated_data(data: dict,
                      ticker: str,
                      sheet_name: str,
                      field: list):
    sorted_data = {}
    date = ''
    balance_sheet = data[ticker][sheet_name]
    for report_date in balance_sheet:
        date = report_date['date']
        sorted_data[date] = report_date[field]
    return sorted_data


#   tuple should contain the sheet_name and the field required
def get_specific_value(data: dict,
                       ticker: str,
                       sheet_name: str,
                       field: str):
    data_sheet = data[ticker][sheet_name]
    return data_sheet[0][field]


#   name should include ticker and date of earnings
def create_all_workbook(name: str):
    file_name = cnst.DATA_PATH + name + '.xlsx'
    wb = Workbook()
    wb.create_sheet(cnst.REPORT)
    wb.remove(wb['Sheet'])
    wb.save(file_name)
    wb.close()


def create_workbook(name: str):
    file_name = cnst.DATA_PATH + name + '.xlsx'
    wb = Workbook()
    wb.create_sheet(cnst.QUARTER)
    wb.create_sheet(cnst.ANNUAL)
    wb.remove(wb['Sheet'])
    wb.save(file_name)
    wb.close()


def autofill_report_xl(data: dict, file_name: str):
    path = cnst.DATA_PATH + file_name + '.xlsx'
    wb = load_workbook(path)
    ws = wb[file_name]
    col_pos = 1
    for col, param in enumerate(cnst.PROFILE_PARAMS):
        if param[0] == 'Description':
            break
        ws.cell(column=col+1, row=1, value=param[0])
        col_pos += 1
    for param in cnst.CORE_PARAMS:
        if param[0] == 'Date':
            continue
        ws.cell(column=col_pos, row=1, value=param[0])
        col_pos += 1
    for row, ticker in enumerate(data):
        for col, param in enumerate(cnst.PROFILE_PARAMS):
            col_pos = col + 1
            if param[0] == 'Description':
                break
            field_val = data[ticker]['profile'][0][param[1]]
            ws.cell(column=col_pos, row=row+2, value=field_val)
        for col, param in enumerate(cnst.CORE_PARAMS):
            if param[0] == 'Date':
                continue
            datasheet = param[2] + '_' + cnst.ANNUAL
            field_val = data[ticker][datasheet][0][param[1]]
            ws.cell(column=col_pos, row=row+2, value=field_val)
            col_pos += 1
    wb.save(path)
    wb.close()


#   fills xl file with data
def autofill_xl(data: dict, ticker: str, file_name: str, period: str):
    path = cnst.DATA_PATH + file_name + '.xlsx'
    wb = load_workbook(path)
    ws = wb[period]
    row_pos = 1
    date_row = 0
    date_list = {}
    for row, param in enumerate(cnst.PROFILE_PARAMS):
        param_val = get_specific_value(data, ticker, 'profile', param[1])
        ws.cell(column=1, row=row+1, value=param[0])
        ws.cell(column=2, row=row+1, value=param_val)
        row_pos += 1
    for row, param in enumerate(cnst.CORE_PARAMS):
        current_row = row_pos + row
        if row == 0:
            date_row = current_row
            ws.cell(column=1, row=current_row, value=param[0])
            datasheet = check_update_datasheet(data, ticker, period)
            for col, date in enumerate(data[ticker][datasheet]):
                ws.cell(column=col+2, row=current_row, value=date['date'])
                date_list[date['date']] = (col+2, current_row)
            continue
        sheet_name = param[2] + '_' + period
        field = param[1]
        data_dates = sorted_dated_data(data, ticker, sheet_name, field)
        ws.cell(column=1, row=current_row, value=param[0])
        for ref_date in date_list:
            if ref_date in data_dates:
                ws.cell(column=date_list[ref_date][0], row=current_row, value=data_dates[ref_date])
    wb.save(path)
    wb.close()


def check_update_datasheet(data: dict, ticker: str, period: str):
    most_recent_date = datetime.datetime.strptime('1950-12-01', "%Y-%m-%d")
    most_recent_datasheet = ''
    ticker_data = data[ticker]
    for datasheet in ticker_data:
        if period not in datasheet:
            continue
        temp = get_specific_value(data, ticker, datasheet, 'date')
        new_time_date = datetime.datetime.strptime(temp, "%Y-%m-%d")
        if most_recent_date < new_time_date:
            most_recent_date = new_time_date
            most_recent_datasheet = datasheet
    return most_recent_datasheet


def color_up(current_value: int, last_value: int):
    try:
        avg = 100 * (current_value - last_value) / abs(last_value)
        style = ''
        if avg >= 100:
            color = Color(rgb='24CA2F')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg >= 80 and avg < 100:
            color = Color(rgb='47D150')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg >= 60 and avg < 80:
            color = Color(rgb='69D871')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg >= 40 and avg < 60:
            color = Color(rgb='8CDE91')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg >= 20 and avg < 40:
            color = Color(rgb='AEE5B2')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg >= 0 and avg < 20:
            color = Color(rgb='D1ECD3')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg < 0 and avg > -20:
            color = Color(rgb='FFA9A7')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg <= -20 and avg > -40:
            color = Color(rgb='FF8786')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg <= -40 and avg > -60:
            color = Color(rgb='FF6564')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg <= -60 and avg > -80:
            color = Color(rgb='FF4443')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg <= -80 and avg > -100:
            color = Color(rgb='FF2221')
            style = PatternFill(fgColor=color, fill_type='solid')
        if avg <= -100:
            color = Color(rgb='FF0000')
            style = PatternFill(fgColor=color, fill_type='solid')
        return style
    except:
        logging.info("Please enter valid values for the parameters")


def prettypy(file_name: str, period: str):
    path = cnst.DATA_PATH + file_name + '.xlsx'
    wb = load_workbook(path)
    ws = wb[period]
    columns_best_fit(ws)
    for row in range(0, 30):
        cell = ws[xlref(row, 0)]
        cell.font = cnst.STYLE_PARAM
    for col in range(1, 60):
        cell = ws[xlref(7, col)]
        cell.font = cnst.STYLE_PARAM
    for row in range(8, 29):
        for col in range(1, 60):
            current_cell = ws[xlref(row, col)]
            next_cell = ws[xlref(row, col+1)]
            style = color_up(current_cell.value, next_cell.value)
            if style:
                current_cell.fill = style
    wb.save(path)
    wb.close()


def prettypy_report(file_name: str):
    path = cnst.DATA_PATH + file_name + '.xlsx'
    wb = load_workbook(path)
    ws = wb[file_name]
    columns_best_fit(ws)
    for col in range(0, 30):
        cell = ws[xlref(0, col)]
        cell.font = cnst.STYLE_PARAM
    for row in range(1, 100):
        cell = ws[xlref(row, 0)]
        cell.font = cnst.STYLE_PARAM
    wb.save(path)
    wb.close()


def columns_best_fit(ws: worksheet.worksheet.Worksheet):
    column_letters = tuple(get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].bestFit = True


def xlref(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def gen_xl(data: dict, only_report=False):
    if not only_report:
        for stonk in data:
            date = data[stonk]['calendar']['date']
            file_name = date + '_' + stonk
            create_workbook(file_name)
            for period in [cnst.ANNUAL, cnst.QUARTER]:
                autofill_xl(data, stonk, file_name, period)
                prettypy(file_name, period)
    create_all_workbook(cnst.REPORT)
    autofill_report_xl(data, cnst.REPORT)
    prettypy_report(cnst.REPORT)


def gen_xl_single(data):
    for stonk in data:
        file_name = stonk
        create_workbook(file_name)
        for period in [cnst.ANNUAL, cnst.QUARTER]:
            autofill_xl(data, stonk, file_name, period)
            prettypy(file_name, period)
